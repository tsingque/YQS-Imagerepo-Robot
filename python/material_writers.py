#!/usr/bin/env python3
"""Writers for YQS-Imagerepo material metadata files."""

from __future__ import annotations

import csv
import re
import subprocess
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


PYTHON_DIR = Path(__file__).resolve().parent
PROJECT_DIR = PYTHON_DIR.parent

MD_TABLE_PATH = PROJECT_DIR / "案例素材清单_表格.md"
RESOLUTION_CSV_PATH = PROJECT_DIR / "素材分辨率.csv"
XLSX_SYNC_SCRIPT = PYTHON_DIR / "md_table_to_material_xlsx.py"
MATERIAL_XLSX_PATH = PROJECT_DIR / "案例素材清单.xlsx"
MAIN_SHEET_NAME = "案例素材清单"

MD_HEADERS = [
    "分类",
    "文件名",
    "案例名称",
    "图片内容",
    "想放在哪（章节/论点）",
    "配图说明文字（图注/要点）",
    "关键数据",
    "来源/版权",
    "状态",
]


def sanitize_cell(value: object) -> str:
    text = "" if value is None else str(value).strip()
    return text.replace("\n", " ").replace("\r", " ").replace("|", "/")


def has_chinese(text: str) -> bool:
    return re.search(r"[\u4e00-\u9fff]", text) is not None


def safe_filename_part(text: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\s]+", "_", text.strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned or "未命名素材"


def ensure_chinese_filename(result: dict) -> str:
    source_file = str(result.get("source_file") or "")
    filename = str(result.get("suggested_filename") or result.get("filename") or source_file or "未命名素材.png")
    suffix = Path(filename).suffix or Path(source_file).suffix or ".png"

    if has_chinese(Path(filename).stem):
        return filename

    case_name = str(result.get("case_name") or result.get("image_content") or result.get("category") or "未命名素材")
    return f"{safe_filename_part(case_name)}{suffix}"


def split_markdown_row(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def ensure_md_table() -> None:
    if MD_TABLE_PATH.is_file() and MD_TABLE_PATH.read_text(encoding="utf-8").strip():
        return

    header = "| " + " | ".join(MD_HEADERS) + " |"
    separator = "|" + "|".join(["---"] * len(MD_HEADERS)) + "|"
    MD_TABLE_PATH.write_text(header + "\n" + separator + "\n", encoding="utf-8")


def existing_md_filenames() -> set[str]:
    if not MD_TABLE_PATH.is_file():
        return set()
    filenames: set[str] = set()
    for line_number, line in enumerate(MD_TABLE_PATH.read_text(encoding="utf-8").splitlines(), start=1):
        if line_number <= 2 or not line.strip().startswith("|"):
            continue
        cells = split_markdown_row(line)
        if len(cells) >= 2:
            filenames.add(cells[1])
    return filenames


def normalize_result(result: dict, fallback_resolution: str) -> dict[str, str]:
    filename = ensure_chinese_filename(result)
    category = result.get("category") or result.get("material_type") or result.get("suggested_folder") or "其他_待判断"
    suggested_folder = result.get("suggested_folder") or category
    source = sanitize_cell(result.get("bitable_source") or result.get("来源") or result.get("source", ""))
    commercial = sanitize_cell(result.get("bitable_commercial") or result.get("是否可商用") or result.get("commercial", ""))
    source_copyright = result.get("source_copyright")
    if source or commercial:
        source_copyright = f"{source or '来源未填'}/{commercial or '无'}"

    return {
        "分类": sanitize_cell(suggested_folder),
        "文件名": sanitize_cell(filename),
        "案例名称": sanitize_cell(result.get("case_name", "")),
        "图片内容": sanitize_cell(result.get("image_content", "")),
        "想放在哪（章节/论点）": sanitize_cell(result.get("ppt_usage", "")),
        "配图说明文字（图注/要点）": sanitize_cell(result.get("caption", "")),
        "关键数据": sanitize_cell(result.get("key_data", "无") or "无"),
        "来源/版权": sanitize_cell(source_copyright or "来源未填/无"),
        "状态": sanitize_cell(result.get("status", "已识别")),
        "resolution": sanitize_cell(result.get("resolution", fallback_resolution)),
    }


def append_record_to_md(record: dict[str, str]) -> bool:
    ensure_md_table()
    if record["文件名"] in existing_md_filenames():
        return False

    row = "| " + " | ".join(record[header] for header in MD_HEADERS) + " |"
    with MD_TABLE_PATH.open("a", encoding="utf-8") as file:
        file.write(row + "\n")
    return True


def upsert_resolution(filename: str, resolution: str) -> None:
    rows: list[dict[str, str]] = []
    if RESOLUTION_CSV_PATH.is_file():
        with RESOLUTION_CSV_PATH.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            for row in reader:
                if row.get("filename") and row.get("filename") != filename:
                    rows.append({"filename": row["filename"], "resolution": row.get("resolution", "")})

    rows.append({"filename": filename, "resolution": resolution})
    with RESOLUTION_CSV_PATH.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=["filename", "resolution"])
        writer.writeheader()
        writer.writerows(rows)


def copy_cell_style(source_cell, target_cell) -> None:
    if not source_cell.has_style:
        return
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def existing_xlsx_filenames() -> set[str]:
    if not MATERIAL_XLSX_PATH.is_file():
        return set()
    workbook = load_workbook(MATERIAL_XLSX_PATH)
    if MAIN_SHEET_NAME not in workbook.sheetnames:
        return set()
    sheet = workbook[MAIN_SHEET_NAME]
    filenames: set[str] = set()
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if value:
            filenames.add(str(value))
    return filenames


def append_record_to_xlsx(record: dict[str, str]) -> bool:
    if not MATERIAL_XLSX_PATH.is_file():
        raise FileNotFoundError(f"未找到 Excel 清单: {MATERIAL_XLSX_PATH}")
    if record["文件名"] in existing_xlsx_filenames():
        return False

    workbook = load_workbook(MATERIAL_XLSX_PATH)
    if MAIN_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Excel 缺少工作表: {MAIN_SHEET_NAME}")

    sheet = workbook[MAIN_SHEET_NAME]
    target_headers = [
        "文件名",
        "案例名称",
        "图片内容",
        "想放在哪（章节/论点）",
        "配图说明文字（图注/要点）",
        "关键数据",
        "来源/版权",
        "状态",
        "存放文件夹",
    ]
    headers = [sheet.cell(1, column).value for column in range(1, len(target_headers) + 1)]
    if headers != target_headers:
        raise ValueError(f"Excel 表头不匹配。期望: {target_headers}; 实际: {headers}")

    values = [
        record["文件名"],
        record["案例名称"],
        record["图片内容"],
        record["想放在哪（章节/论点）"],
        record["配图说明文字（图注/要点）"],
        record["关键数据"],
        record["来源/版权"],
        record["状态"],
        record["分类"],
    ]

    source_row = sheet.max_row if sheet.max_row >= 2 else 1
    target_row = sheet.max_row + 1
    if sheet.row_dimensions[source_row].height is not None:
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column, value in enumerate(values, start=1):
        copy_cell_style(sheet.cell(source_row, column), sheet.cell(target_row, column))
        sheet.cell(target_row, column).value = value

    workbook.save(MATERIAL_XLSX_PATH)
    return True


def sync_md_to_xlsx() -> None:
    subprocess.run(
        ["python3", str(XLSX_SYNC_SCRIPT)],
        cwd=str(PROJECT_DIR),
        check=True,
    )


def write_recognition_result(result: dict, fallback_resolution: str) -> tuple[dict[str, str], bool]:
    record = normalize_result(result, fallback_resolution)
    appended_md = append_record_to_md(record)
    appended_xlsx = append_record_to_xlsx(record)
    upsert_resolution(record["文件名"], record["resolution"])
    return record, appended_md or appended_xlsx
