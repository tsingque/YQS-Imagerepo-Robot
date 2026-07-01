#!/usr/bin/env python3
"""Archive material images and reset local/Feishu material records."""

from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from feishu_bitable_client import FeishuBitableClient
from material_writers import MD_HEADERS


PYTHON_DIR = Path(__file__).resolve().parent
PROJECT_DIR = PYTHON_DIR.parent
DESKTOP_DIR = Path.home() / "Desktop"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff", ".heic", ".heif"}
MATERIAL_IMAGE_ROOTS = [
    PROJECT_DIR / "case_materials",
    PROJECT_DIR / "image_compressor" / "images_raw",
    PROJECT_DIR / "image_compressor" / "images_compressed",
    PROJECT_DIR / "image_compressor" / "images_recognized",
    PROJECT_DIR / "deer-flow" / "backend" / ".deer-flow",
]
MD_TABLE_PATH = PROJECT_DIR / "案例素材清单_表格.md"
RESOLUTION_CSV_PATH = PROJECT_DIR / "素材分辨率.csv"
MATERIAL_XLSX_PATH = PROJECT_DIR / "案例素材清单.xlsx"
RUNTIME_FILES_TO_CLEAR = [
    PROJECT_DIR / "runtime" / "bitable_file_map.json",
    PROJECT_DIR / "runtime" / "bitable_records.json",
    PROJECT_DIR / "runtime" / "bitable_sync_state.json",
    PROJECT_DIR / "runtime" / "processed_files.json",
    PROJECT_DIR / "runtime" / "recognition_state.json",
    PROJECT_DIR / "runtime" / "ai_token_usage.json",
    PROJECT_DIR / "runtime" / "deerflow_agent_state.json",
]


def is_image(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS


def archive_material_images() -> tuple[Path, int]:
    archive_dir = DESKTOP_DIR / f"YQS素材图片归档_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    archive_dir.mkdir(parents=True, exist_ok=True)
    moved = 0
    for root in MATERIAL_IMAGE_ROOTS:
        if not root.exists():
            continue
        for source in sorted(root.rglob("*")):
            if not is_image(source):
                continue
            relative = source.relative_to(PROJECT_DIR)
            target = archive_dir / relative
            target.parent.mkdir(parents=True, exist_ok=True)
            if target.exists():
                target = target.with_name(f"{target.stem}_{moved}{target.suffix}")
            shutil.move(str(source), str(target))
            moved += 1
    return archive_dir, moved


def clear_markdown_table() -> None:
    header = "| " + " | ".join(MD_HEADERS) + " |"
    separator = "|" + "|".join(["---"] * len(MD_HEADERS)) + "|"
    MD_TABLE_PATH.write_text(header + "\n" + separator + "\n", encoding="utf-8")


def clear_resolution_csv() -> None:
    RESOLUTION_CSV_PATH.write_text("filename,resolution\n", encoding="utf-8")


def clear_workbook_data() -> None:
    if not MATERIAL_XLSX_PATH.exists():
        return
    workbook = load_workbook(MATERIAL_XLSX_PATH)
    for sheet in workbook.worksheets:
        if sheet.max_row <= 1:
            continue
        sheet.delete_rows(2, sheet.max_row - 1)
    workbook.save(MATERIAL_XLSX_PATH)


def clear_runtime_files() -> None:
    for path in RUNTIME_FILES_TO_CLEAR:
        if not path.exists():
            continue
        if path.name.endswith("_state.json") or path.name == "deerflow_agent_state.json":
            payload = {
                "running": False,
                "message": "已清空素材状态。",
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "summary": {},
                "errors": [],
            }
        else:
            payload = {}
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def clear_bitable_records() -> int:
    client = FeishuBitableClient()
    if not client.configured():
        return 0
    records = client.list_records(page_size=500)
    deleted = 0
    for record in records:
        record_id = str(record.get("record_id") or record.get("id") or "").strip()
        if not record_id:
            continue
        client.request("DELETE", f"/bitable/v1/apps/{client.app_token}/tables/{client.table_id}/records/{record_id}")
        deleted += 1
    return deleted


def main() -> None:
    archive_dir, moved_images = archive_material_images()
    clear_markdown_table()
    clear_resolution_csv()
    clear_workbook_data()
    clear_runtime_files()
    deleted_records = clear_bitable_records()
    print(json.dumps({
        "ok": True,
        "archive_dir": str(archive_dir),
        "moved_images": moved_images,
        "cleared_files": [
            str(MD_TABLE_PATH.relative_to(PROJECT_DIR)),
            str(RESOLUTION_CSV_PATH.relative_to(PROJECT_DIR)),
            str(MATERIAL_XLSX_PATH.relative_to(PROJECT_DIR)),
        ],
        "deleted_bitable_records": deleted_records,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
