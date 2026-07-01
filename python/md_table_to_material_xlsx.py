#!/usr/bin/env python3
"""
Convert the standalone Markdown material table into a timestamped temp workbook,
append its rows to 案例素材清单.xlsx, then archive the temp workbook in md_trans_repo/.

Run from the project root:
  python3 python/md_table_to_material_xlsx.py

Preview without modifying the target workbook:
  python3 python/md_table_to_material_xlsx.py --dry-run
"""

from __future__ import annotations

import argparse
import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


PYTHON_DIR = Path(__file__).resolve().parent
PROJECT_DIR = PYTHON_DIR.parent

DEFAULT_MD_PATH = PROJECT_DIR / "案例素材清单_表格.md"
DEFAULT_TARGET_XLSX = PROJECT_DIR / "案例素材清单.xlsx"
DEFAULT_ARCHIVE_DIR = PROJECT_DIR / "md_trans_repo"
MAIN_SHEET_NAME = "案例素材清单"

TARGET_HEADERS = [
    "文件名",
    "案例名称",
    "图片内容",
    "想放在哪（章节/论点）",
    "配图说明文字（图注/要点）",
    "关键数据",
    "来源/版权",
    "状态",
    "存放文件夹",
]

MD_HEADERS = [
    "分类",
    "文件名",
    "案例名称",
    "图片内容",
    "想放在哪（章节/论点）",
    "配图说明文字（图注/要点）",
    "关键数据",
    "来源/版权",
    "状态",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert 案例素材清单_表格.md into a temp workbook and append rows to 案例素材清单.xlsx."
    )
    parser.add_argument("--md", default=str(DEFAULT_MD_PATH), help="Markdown table file path.")
    parser.add_argument("--target", default=str(DEFAULT_TARGET_XLSX), help="Target material workbook path.")
    parser.add_argument("--archive-dir", default=str(DEFAULT_ARCHIVE_DIR), help="Directory for archived temp workbooks.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and report only; do not write files.")
    return parser.parse_args()


def is_separator_row(cells: list[str]) -> bool:
    return all(re.fullmatch(r":?-{3,}:?", cell.strip()) for cell in cells)


def split_markdown_row(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def parse_markdown_table(md_path: Path) -> list[dict[str, str]]:
    if not md_path.is_file():
        raise FileNotFoundError(f"未找到 Markdown 表格文件: {md_path}")

    table_rows: list[list[str]] = []
    for line in md_path.read_text(encoding="utf-8").splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = split_markdown_row(line)
        if not cells or is_separator_row(cells):
            continue
        table_rows.append(cells)

    if not table_rows:
        raise ValueError(f"未在 {md_path.name} 中找到 Markdown 表格。")

    headers = table_rows[0]
    if headers != MD_HEADERS:
        raise ValueError(f"Markdown 表头不匹配。期望: {MD_HEADERS}; 实际: {headers}")

    records: list[dict[str, str]] = []
    for row_number, cells in enumerate(table_rows[1:], start=2):
        if len(cells) != len(MD_HEADERS):
            raise ValueError(f"第 {row_number} 行列数不匹配: 期望 {len(MD_HEADERS)} 列，实际 {len(cells)} 列。")
        records.append(dict(zip(MD_HEADERS, cells)))

    return records


def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(PROJECT_DIR))
    except ValueError:
        return str(path)


def map_record_to_target_row(record: dict[str, str]) -> list[str]:
    return [
        record["文件名"],
        record["案例名称"],
        record["图片内容"],
        record["想放在哪（章节/论点）"],
        record["配图说明文字（图注/要点）"],
        record["关键数据"],
        record["来源/版权"],
        record["状态"],
        record["分类"],
    ]


def copy_cell_style(source_cell, target_cell) -> None:
    if not source_cell.has_style:
        return
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def copy_row_style(sheet: Worksheet, source_row: int, target_row: int, max_column: int) -> None:
    source_height = sheet.row_dimensions[source_row].height
    if source_height is not None:
        sheet.row_dimensions[target_row].height = source_height

    for column in range(1, max_column + 1):
        copy_cell_style(sheet.cell(source_row, column), sheet.cell(target_row, column))


def snapshot_row_style(sheet: Worksheet, row: int, max_column: int) -> dict:
    return {
        "height": sheet.row_dimensions[row].height,
        "cells": [
            {
                "font": copy(sheet.cell(row, column).font),
                "fill": copy(sheet.cell(row, column).fill),
                "border": copy(sheet.cell(row, column).border),
                "alignment": copy(sheet.cell(row, column).alignment),
                "number_format": sheet.cell(row, column).number_format,
                "protection": copy(sheet.cell(row, column).protection),
            }
            for column in range(1, max_column + 1)
        ],
    }


def apply_row_style_snapshot(sheet: Worksheet, snapshot: dict, target_row: int) -> None:
    if snapshot["height"] is not None:
        sheet.row_dimensions[target_row].height = snapshot["height"]

    for column, style in enumerate(snapshot["cells"], start=1):
        cell = sheet.cell(target_row, column)
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.number_format = style["number_format"]
        cell.protection = copy(style["protection"])


def validate_target_sheet(sheet: Worksheet) -> None:
    headers = [sheet.cell(1, column).value for column in range(1, len(TARGET_HEADERS) + 1)]
    if headers != TARGET_HEADERS:
        raise ValueError(f"目标 Excel 表头不匹配。期望: {TARGET_HEADERS}; 实际: {headers}")


def clear_data_rows(sheet: Worksheet) -> None:
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)


def append_rows(
    sheet: Worksheet,
    rows: list[list[str]],
    style_source_row: int | None = None,
    style_snapshot: dict | None = None,
) -> None:
    start_row = sheet.max_row + 1
    style_row = style_source_row or max(1, sheet.max_row)

    for row_offset, values in enumerate(rows):
        target_row = start_row + row_offset
        if style_snapshot is not None:
            apply_row_style_snapshot(sheet, style_snapshot, target_row)
        elif style_row > 0:
            copy_row_style(sheet, style_row, target_row, len(TARGET_HEADERS))
        for column, value in enumerate(values, start=1):
            sheet.cell(target_row, column).value = value


def build_temp_workbook(target_xlsx: Path, rows: list[list[str]], temp_path: Path) -> None:
    workbook = load_workbook(target_xlsx)
    if MAIN_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"目标 Excel 缺少工作表: {MAIN_SHEET_NAME}")

    sheet = workbook[MAIN_SHEET_NAME]
    validate_target_sheet(sheet)
    style_snapshot = snapshot_row_style(sheet, 2 if sheet.max_row >= 2 else 1, len(TARGET_HEADERS))
    clear_data_rows(sheet)
    append_rows(sheet, rows, style_snapshot=style_snapshot)
    workbook.save(temp_path)


def append_temp_to_target(target_xlsx: Path, temp_path: Path) -> int:
    target_workbook = load_workbook(target_xlsx)
    temp_workbook = load_workbook(temp_path)

    if MAIN_SHEET_NAME not in target_workbook.sheetnames:
        raise ValueError(f"目标 Excel 缺少工作表: {MAIN_SHEET_NAME}")
    if MAIN_SHEET_NAME not in temp_workbook.sheetnames:
        raise ValueError(f"temp Excel 缺少工作表: {MAIN_SHEET_NAME}")

    target_sheet = target_workbook[MAIN_SHEET_NAME]
    temp_sheet = temp_workbook[MAIN_SHEET_NAME]
    validate_target_sheet(target_sheet)
    validate_target_sheet(temp_sheet)

    rows: list[list[str]] = []
    for row in temp_sheet.iter_rows(min_row=2, max_col=len(TARGET_HEADERS), values_only=True):
        if any(value not in (None, "") for value in row):
            rows.append(["" if value is None else str(value) for value in row])

    if not rows:
        return 0

    style_source_row = target_sheet.max_row if target_sheet.max_row >= 2 else 1
    append_rows(target_sheet, rows, style_source_row=style_source_row)
    target_workbook.save(target_xlsx)
    return len(rows)


def main() -> None:
    args = parse_args()
    md_path = Path(args.md).resolve()
    target_xlsx = Path(args.target).resolve()
    archive_dir = Path(args.archive_dir).resolve()

    if not target_xlsx.is_file():
        raise FileNotFoundError(f"未找到目标 Excel: {target_xlsx}")

    records = parse_markdown_table(md_path)
    rows = [map_record_to_target_row(record) for record in records]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_name = f"{timestamp}_temp.xlsx"
    temp_path = PROJECT_DIR / temp_name
    archived_temp_path = archive_dir / temp_name

    print(f"Markdown 表格: {display_path(md_path)}")
    print(f"目标 Excel: {display_path(target_xlsx)}")
    print(f"待追加行数: {len(rows)}")
    print(f"temp 文件: {temp_name}")

    if args.dry_run:
        print("dry-run: 不写入 Excel，不移动 temp 文件。")
        return

    archive_dir.mkdir(parents=True, exist_ok=True)
    build_temp_workbook(target_xlsx, rows, temp_path)
    appended_count = append_temp_to_target(target_xlsx, temp_path)
    shutil.move(str(temp_path), str(archived_temp_path))

    print(f"已追加行数: {appended_count}")
    print(f"temp 已归档: {display_path(archived_temp_path)}")


if __name__ == "__main__":
    main()
