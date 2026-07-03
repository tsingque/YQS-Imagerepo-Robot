import shutil
import subprocess
import tempfile
import unittest
from pathlib import Path
from typing import Union

from openpyxl import Workbook, load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
SCRIPT_PATH = PROJECT_DIR / "scripts" / "clear_project_environment.sh"


TARGET_HEADERS = [
    "文件名",
    "案例名称",
    "图片内容",
    "想放在哪（章节/论点）",
    "配图说明文字（图注/要点）",
    "关键数据",
    "来源/版权",
    "状态",
    "存放文件夹",
]


class ClearProjectEnvironmentScriptTests(unittest.TestCase):
    def make_project(self) -> Path:
        root = Path(tempfile.mkdtemp())
        scripts_dir = root / "scripts"
        scripts_dir.mkdir()
        shutil.copy2(SCRIPT_PATH, scripts_dir / SCRIPT_PATH.name)
        return root

    def run_script(self, root: Path, *args: str) -> subprocess.CompletedProcess[str]:
        env = {
            "PATH": "/usr/bin:/bin:/usr/local/bin",
            "YQS_CLEAN_TIMESTAMP": "20260101_000000",
        }
        return subprocess.run(
            ["bash", str(root / "scripts" / "clear_project_environment.sh"), *args],
            cwd=str(root),
            env=env,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=True,
        )

    def write(self, root: Path, rel: str, content: Union[bytes, str] = "data") -> Path:
        path = root / rel
        path.parent.mkdir(parents=True, exist_ok=True)
        if isinstance(content, bytes):
            path.write_bytes(content)
        else:
            path.write_text(content, encoding="utf-8")
        return path

    def make_workbook(self, root: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "案例素材清单"
        sheet.append(TARGET_HEADERS)
        sheet.append(["demo.png", "demo", "old", "", "", "", "", "已识别", "通用素材库"])
        workbook.save(root / "案例素材清单.xlsx")

    def test_default_cleanup_clears_workflow_data_and_keeps_env(self):
        root = self.make_project()
        try:
            self.write(root, "runtime/state.json", "{}")
            self.write(root, "case_materials/通用素材库/demo.png", b"image")
            self.write(root, "image_compressor/images_raw/raw.png", b"raw")
            self.write(root, "image_compressor/compression_report.md", "old report")
            self.write(root, "案例素材清单_表格.md", "| old |\n")
            self.write(root, "素材分辨率.csv", "filename,resolution\ndemo.png,1x1\n")
            self.make_workbook(root)
            self.write(root, ".env", "SECRET=local")
            self.write(root, "config.yaml", "config: local")

            result = self.run_script(root, "--yes")

            self.assertIn("Cleanup complete", result.stdout)
            backup = root / ".cleanup_backups" / "20260101_000000"
            self.assertTrue((backup / "runtime/state.json").is_file())
            self.assertTrue((backup / "case_materials/通用素材库/demo.png").is_file())
            self.assertTrue((backup / "image_compressor/compression_report.md").is_file())
            self.assertTrue((backup / "案例素材清单.xlsx").is_file())
            self.assertTrue((root / "runtime").is_dir())
            self.assertFalse((root / "runtime/state.json").exists())
            self.assertTrue((root / "case_materials").is_dir())
            self.assertFalse((root / "case_materials/通用素材库/demo.png").exists())
            self.assertEqual((root / "素材分辨率.csv").read_text(encoding="utf-8"), "filename,resolution\n")
            self.assertNotIn("old", (root / "案例素材清单_表格.md").read_text(encoding="utf-8"))
            workbook = load_workbook(root / "案例素材清单.xlsx")
            self.assertEqual(workbook["案例素材清单"].max_row, 1)
            self.assertTrue((root / ".env").is_file())
            self.assertTrue((root / "config.yaml").is_file())
        finally:
            shutil.rmtree(root)

    def test_include_env_moves_environment_config_but_keeps_examples(self):
        root = self.make_project()
        try:
            self.write(root, ".env", "SECRET=local")
            self.write(root, ".env.example", "SECRET=example")
            self.write(root, "config.yaml", "config: local")
            self.write(root, "deer-flow/config.yaml", "config: deer")
            self.write(root, "deer-flow/extensions_config.json", "{}")
            self.write(root, "deer-flow/frontend/.env", "FRONTEND=local")

            self.run_script(root, "--yes", "--include-env")

            backup = root / ".cleanup_backups" / "20260101_000000"
            self.assertFalse((root / ".env").exists())
            self.assertFalse((root / "config.yaml").exists())
            self.assertFalse((root / "deer-flow/config.yaml").exists())
            self.assertFalse((root / "deer-flow/extensions_config.json").exists())
            self.assertFalse((root / "deer-flow/frontend/.env").exists())
            self.assertTrue((root / ".env.example").is_file())
            self.assertTrue((backup / ".env").is_file())
            self.assertTrue((backup / "config.yaml").is_file())
            self.assertTrue((backup / "deer-flow/config.yaml").is_file())
        finally:
            shutil.rmtree(root)

    def test_dry_run_does_not_change_files(self):
        root = self.make_project()
        try:
            self.write(root, "runtime/state.json", "{}")
            self.write(root, "case_materials/demo.png", b"image")
            self.write(root, ".env", "SECRET=local")

            result = self.run_script(root, "--dry-run", "--yes", "--include-env")

            self.assertIn("Dry-run complete", result.stdout)
            self.assertTrue((root / "runtime/state.json").is_file())
            self.assertTrue((root / "case_materials/demo.png").is_file())
            self.assertTrue((root / ".env").is_file())
            self.assertFalse((root / ".cleanup_backups").exists())
        finally:
            shutil.rmtree(root)


if __name__ == "__main__":
    unittest.main()
